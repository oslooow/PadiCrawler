import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import json
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from compare_scraper import fetch_toko_products, compare_multiple_products, cleanup
import os
from pathlib import Path

class CompareBarangApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Compare Toko")
        self.root.geometry("800x600")

        # Configure styles
        self.style = ttk.Style()
        self.configure_styles()

        # Main container with padding
        self.main_container = ttk.Frame(root, padding="20")
        self.main_container.pack(fill=tk.BOTH, expand=False)

        # Configure grid
        self.main_container.columnconfigure(0, weight=1)

        # Header
        self.header = ttk.Label(
            self.main_container,
            text="Compare Toko",
            style="Header.TLabel"
        )
        self.header.pack(pady=(0, 30))

        # Input Container Frame (for centering inputs)
        self.input_container = ttk.Frame(self.main_container)
        self.input_container.pack(fill=tk.X, padx=20)

        # Toko 1 Frame
        self.toko1_frame = ttk.Frame(self.input_container)
        self.toko1_frame.pack(fill=tk.X, pady=(0, 20))
        
        self.toko1_label = ttk.Label(
            self.toko1_frame,
            text="Your Toko Code:",
            style="Label.TLabel",
            anchor="center"
        )
        self.toko1_label.pack(fill=tk.X)
        
        self.toko1_entry = ttk.Entry(self.toko1_frame, width=40, justify='center')
        self.toko1_entry.pack(pady=(5, 0))

        # Toko 2 Frame
        self.toko2_frame = ttk.Frame(self.input_container)
        self.toko2_frame.pack(fill=tk.X, pady=(0, 20))
        
        self.toko2_label = ttk.Label(
            self.toko2_frame,
            text="Compared Toko (max 3, separate with ;):",
            style="Label.TLabel",
            anchor="center"
        )
        self.toko2_label.pack(fill=tk.X)
        
        self.toko2_entry = ttk.Entry(self.toko2_frame, width=40, justify='center')
        self.toko2_entry.pack(pady=(5, 0))

        # Status Label
        self.status_label = ttk.Label(
            self.main_container,
            text="",
            style="Status.TLabel"
        )
        self.status_label.pack(pady=5)

        # Progress Frame
        self.progress_frame = ttk.Frame(self.main_container)
        self.progress_frame.pack(fill=tk.X, pady=10)
        
        # Progress Bar
        self.progress_bar = ttk.Progressbar(
            self.progress_frame,
            mode='indeterminate',
            style="Accent.Horizontal.TProgressbar"
        )

        # Compare Button
        self.compare_button = ttk.Button(
            self.main_container,
            text="Compare Toko",
            command=self.compare_toko,
            style="Compare.TButton"
        )
        self.compare_button.pack(pady=20)

        # Results Frame
        self.results_frame = ttk.Frame(self.main_container)
        self.results_frame.pack(fill=tk.BOTH, expand=True)

        # Results Text Widget
        self.results_text = tk.Text(
            self.results_frame,
            height=15,
            width=50,
            wrap=tk.WORD,
            font=("Consolas", 10)
        )
        self.results_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Scrollbar for Results
        self.scrollbar = ttk.Scrollbar(
            self.results_frame,
            orient=tk.VERTICAL,
            command=self.results_text.yview
        )
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.results_text.config(yscrollcommand=self.scrollbar.set)

        # Bind window close event
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def on_closing(self):
        """Handle window closing"""
        cleanup()  # Clean up webdriver
        self.root.destroy()

    def configure_styles(self):
        """Configure custom styles"""
        # Configure modern color scheme
        bg_color = "#ffffff"  # Clean white background
        accent_color = "#3498db"  # Nice blue color
        text_color = "#2c3e50"  # Dark blue-gray color
        
        # Frame Styles
        self.style.configure("TFrame", background=bg_color)
        
        # Label Styles
        self.style.configure(
            "Header.TLabel",
            font=("Segoe UI", 24, "bold"),
            foreground=text_color,
            background=bg_color,
            justify="center"
        )
        
        self.style.configure(
            "Label.TLabel",
            font=("Segoe UI", 12),
            background=bg_color,
            foreground=text_color,
            justify="center"
        )
        
        self.style.configure(
            "Status.TLabel",
            font=("Segoe UI", 10),
            background=bg_color,
            foreground=accent_color,
            justify="center"
        )
        
        # Entry Style
        self.style.configure(
            "TEntry",
            fieldbackground="white",
            borderwidth=1,
            relief="solid"
        )
        
        # Button Style
        self.style.configure(
            "Compare.TButton",
            font=("Segoe UI", 12, "bold"),
            background=accent_color,
            foreground="white",
            padding=(20, 10)
        )
        
        # Progress Bar Style
        self.style.configure(
            "Accent.Horizontal.TProgressbar",
            background=accent_color,
            troughcolor="#f0f0f0"
        )

    def compare_toko(self):
        """Compare multiple toko"""
        toko1_code = self.toko1_entry.get().strip()
        toko2_codes = [code.strip() for code in self.toko2_entry.get().split(';') if code.strip()]

        if not toko1_code or not toko2_codes:
            messagebox.showerror("Error", "Please enter both toko codes")
            return

        if len(toko2_codes) > 3:
            messagebox.showerror("Error", "Maximum 3 comparison toko allowed")
            return

        # Show progress bar and update status
        self.progress_bar.pack(fill=tk.X, padx=5)
        self.progress_bar.start()
        self.compare_button.state(['disabled'])
        
        # Clear previous results
        self.results_text.delete(1.0, tk.END)
        
        try:
            # Fetch products from first store
            self.status_label.config(text="Scanning Toko 1 products...")
            toko1_name, toko1_products = fetch_toko_products(toko1_code)
            self.results_text.insert(tk.END, f"Found {len(toko1_products)} products in {toko1_name}\n\n")
            
            # Fetch products from comparison stores
            toko2_list = []
            for code in toko2_codes:
                self.status_label.config(text=f"Scanning comparison toko {code}...")
                name, products = fetch_toko_products(code)
                toko2_list.append((name, products))
                self.results_text.insert(tk.END, f"Found {len(products)} products in {name}\n")
            
            # Compare products
            self.status_label.config(text="Comparing products...")
            unique_products = compare_multiple_products(toko1_products, toko2_list)
            
            # Display and save results
            self.update_results(toko1_name, toko2_list, unique_products)
            self.save_results(toko1_name, toko2_list, unique_products)
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
        finally:
            # Reset UI
            self.progress_bar.stop()
            self.progress_bar.pack_forget()
            self.compare_button.state(['!disabled'])
            self.status_label.config(text="")

    def update_results(self, toko1_name, toko2_list, unique_products):
        """Update the results text widget with comparison results"""
        self.results_text.insert(tk.END, f"\nProducts not found in {toko1_name}:\n\n")
        
        if not unique_products:
            self.results_text.insert(tk.END, "No unique products found.\n")
            return
            
        for product in unique_products:
            self.results_text.insert(tk.END, f"ID: {product['id_barang']}\n")
            self.results_text.insert(tk.END, f"Name: {product['name']}\n")
            self.results_text.insert(tk.END, f"Price: Rp {product['price']:,.0f}\n")
            self.results_text.insert(tk.END, f"Store: {product['source_store']}\n")
            self.results_text.insert(tk.END, "-" * 50 + "\n\n")
            
        self.results_text.insert(tk.END, f"\nComparison completed at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    def save_results(self, toko1_name, toko2_list, unique_products):
        """Save comparison results to Excel file with user-selected location"""
        # Create default directory in Documents
        default_dir = os.path.join(str(Path.home()), "Documents", "CompareResult")
        os.makedirs(default_dir, exist_ok=True)

        # Default filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"comparison_results_{timestamp}.xlsx"
        
        # Ask user for save location
        filename = filedialog.asksaveasfilename(
            initialdir=default_dir,
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Comparison Results As"
        )
        
        if not filename:
            self.results_text.insert(tk.END, "\nFile save cancelled.")
            return
        
        # Create DataFrame with numbered index starting from 1
        df_data = []
        for idx, product in enumerate(unique_products, 1):
            df_data.append({
                'No': idx,
                'Nama Barang': product['name'],
                'Harga': f"Rp {product['price']:,.0f}",
                'Nama Toko': product['source_store']
            })
            
        df = pd.DataFrame(df_data)
        
        # Add header information
        comparison_stores = ", ".join(store[0] for store in toko2_list)
        header_df = pd.DataFrame([
            ['Comparison Results'],
            [f'Toko 1: {toko1_name}'],
            [f'Comparison Toko: {comparison_stores}'],
            [f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'],
            [''],  # Empty row before data
        ])
        
        # Save to Excel
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Write header
            header_df.to_excel(writer, index=False, header=False, sheet_name='Sheet1')
            
            # Write main data starting from row 6
            df.to_excel(writer, index=False, header=True, sheet_name='Sheet1', startrow=5)
            
            # Get the workbook to apply styling
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            # Style the header
            for row in range(1, 5):
                cell = worksheet.cell(row=row, column=1)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left')
            
            # Style the table headers
            header_row = 6
            for col in range(1, 5):
                cell = worksheet.cell(row=header_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 5   # No
            worksheet.column_dimensions['B'].width = 40  # Nama Barang
            worksheet.column_dimensions['C'].width = 15  # Harga
            worksheet.column_dimensions['D'].width = 30  # Nama Toko
        
        self.results_text.insert(tk.END, f"\nResults saved to {filename}")

def main():
    root = tk.Tk()
    app = CompareBarangApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
