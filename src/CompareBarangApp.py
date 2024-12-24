import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import json
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from compare_scraper import fetch_toko_products, compare_multiple_products, cleanup
import os
from pathlib import Path

class CompareBarangApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Compare Toko")
        self.root.geometry("800x700")
        
        # Configure color scheme
        self.bg_color = "#ffffff"  # Clean white background
        self.accent_color = "#3498db"  # Nice blue color
        self.text_color = "#2c3e50"  # Dark blue-gray color
        
        # Configure root window
        self.root.configure(bg=self.bg_color)
        
        # Main container with padding
        self.main_container = tk.Frame(root, bg=self.bg_color, padx=40, pady=30)
        self.main_container.pack(fill=tk.BOTH, expand=True)
        
        # Title Label
        self.title_label = tk.Label(
            self.main_container, 
            text="Compare Toko", 
            font=("Segoe UI", 24, "bold"), 
            bg=self.bg_color, 
            fg=self.text_color
        )
        self.title_label.pack(pady=(0, 30))

        # Input styles
        entry_style = {
            "font": ("Segoe UI", 12),
            "bg": "white",
            "fg": self.text_color,
            "relief": "solid",
            "bd": 1
        }
        
        label_style = {
            "font": ("Segoe UI", 12),
            "bg": self.bg_color,
            "fg": self.text_color
        }
        
        # Toko 1 Frame
        self.toko1_frame = tk.Frame(self.main_container, bg=self.bg_color)
        self.toko1_frame.pack(fill=tk.X, pady=(0, 20))
        
        self.toko1_label = tk.Label(
            self.toko1_frame,
            text="Your Toko Code",
            **label_style
        )
        self.toko1_label.pack(anchor="w")
        
        self.toko1_entry = tk.Entry(
            self.toko1_frame,
            width=50,
            **entry_style
        )
        self.toko1_entry.pack(fill=tk.X, pady=(5, 0))
        
        # Toko 2 Frame
        self.toko2_frame = tk.Frame(self.main_container, bg=self.bg_color)
        self.toko2_frame.pack(fill=tk.X, pady=(0, 20))
        
        self.toko2_label = tk.Label(
            self.toko2_frame,
            text="Compared Toko (max 3, separate with ;)",
            **label_style
        )
        self.toko2_label.pack(anchor="w")
        
        self.toko2_entry = tk.Entry(
            self.toko2_frame,
            width=50,
            **entry_style
        )
        self.toko2_entry.pack(fill=tk.X, pady=(5, 0))
        
        # Compare Button
        self.compare_button = tk.Button(
            self.main_container,
            text="Compare Toko",
            bg=self.accent_color,
            fg="white",
            font=("Segoe UI", 12),
            width=20,
            height=2,
            cursor="hand2",
            bd=0,
            activebackground="#2980b9",
            command=self.compare_toko
        )
        self.compare_button.pack(pady=20)
        
        # Status Label
        self.status_label = tk.Label(
            self.main_container,
            text="",
            font=("Segoe UI", 10),
            bg=self.bg_color,
            fg=self.accent_color
        )
        self.status_label.pack(pady=5)
        
        # Progress Frame
        self.progress_frame = tk.Frame(self.main_container, bg=self.bg_color)
        self.progress_frame.pack(fill=tk.X, pady=10)
        
        # Progress Bar
        self.progress_bar = ttk.Progressbar(
            self.progress_frame,
            orient="horizontal",
            mode="indeterminate",
            style="Accent.Horizontal.TProgressbar"
        )
        
        # Results Frame
        self.results_frame = tk.Frame(self.main_container, bg=self.bg_color, pady=20)
        self.results_frame.pack(fill=tk.BOTH, expand=True)
        
        # Results Text
        self.results_text = tk.Text(
            self.results_frame,
            height=10,
            width=50,
            font=("Consolas", 10),
            bg="#2c3e50",
            fg="#ecf0f1",
            wrap='word'
        )
        self.results_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Scrollbar for Results
        self.scrollbar = ttk.Scrollbar(
            self.results_frame,
            orient=tk.VERTICAL,
            command=self.results_text.yview
        )
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.results_text.config(yscrollcommand=self.scrollbar.set)
        
        # Add hover effect
        self.add_button_hover(self.compare_button)
        
        # Center the window
        self.center_window()

    def add_button_hover(self, button):
        """Add hover effect to buttons"""
        orig_color = button.cget("background")
        darker_color = button.cget("activebackground")

        def on_enter(e):
            button.config(background=darker_color)

        def on_leave(e):
            button.config(background=orig_color)

        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)

    def center_window(self):
        """Center the window on the screen"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")

    def compare_toko(self):
        """Compare multiple toko"""
        toko1_code = self.toko1_entry.get().strip()
        toko2_codes = [code.strip() for code in self.toko2_entry.get().split(';') if code.strip()]

        if not toko1_code:
            messagebox.showerror("Error", "Please enter Toko 1 code")
            return

        if len(toko2_codes) > 3:
            messagebox.showerror("Error", "Maximum 3 comparison toko allowed")
            return

        # Show progress bar and update status
        self.progress_bar.pack(fill=tk.X, padx=5)
        self.progress_bar.start()
        self.compare_button.config(state='disabled')
        
        # Clear previous results
        self.results_text.delete(1.0, tk.END)
        
        try:
            # Fetch products from first store
            self.status_label.config(text="Scanning Toko 1 products...")
            toko1_name, toko1_products = fetch_toko_products(toko1_code)
            self.results_text.insert(tk.END, f"Found {len(toko1_products)} products in {toko1_name}\n\n")
            
            if not toko2_codes:
                # If no comparison toko provided, show all products from toko1
                self.update_results(toko1_name, [], toko1_products, show_all_toko1=True)
                self.save_results(toko1_name, [], toko1_products, show_all_toko1=True)
                return
                
            # Fetch products from comparison stores
            toko2_list = []
            for code in toko2_codes:
                self.status_label.config(text=f"Scanning comparison toko {code}...")
                name, products = fetch_toko_products(code)
                toko2_list.append((name, products))
                self.results_text.insert(tk.END, f"Found {len(products)} products in {name}\n")
            
            # Compare products
            self.status_label.config(text="Comparing products...")
            unique_products = compare_multiple_products(toko1_products, toko2_list)
            
            # Display and save results
            self.update_results(toko1_name, toko2_list, unique_products)
            self.save_results(toko1_name, toko2_list, unique_products)
            
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
        finally:
            # Reset UI
            self.progress_bar.stop()
            self.progress_bar.pack_forget()
            self.compare_button.config(state='normal')
            self.status_label.config(text="")

    def extract_kode_barang(self ,nama_barang):
        """Extract Kode Barang from product name"""
        # Try to find a number at the start of the string
        import re
        match = re.match(r'^(\d+)', nama_barang)
        if match:
            return match.group(1)
        return "0"  # Return "0" if no code found 

    def update_results(self, toko1_name, toko2_list, unique_products, show_all_toko1=False):
        """Update the results text widget with comparison results"""
        if show_all_toko1:
            self.results_text.insert(tk.END, f"\nAll products from {toko1_name}:\n\n")
        else:
            self.results_text.insert(tk.END, f"\nProducts not found in {toko1_name}:\n\n")
        
        if not unique_products:
            self.results_text.insert(tk.END, "No products found.\n")
            return
            
        for product in unique_products:
            nama_barang = product['name']
            kode_barang = self.extract_kode_barang(nama_barang)
            # Remove kode_barang from nama_barang if it exists
            if kode_barang != "0":
                nama_barang = nama_barang[len(kode_barang):].strip(" -;")
            
            self.results_text.insert(tk.END, f"Kode Barang: {kode_barang}\n")
            self.results_text.insert(tk.END, f"Nama Barang: {nama_barang}\n")
            self.results_text.insert(tk.END, f"Harga: Rp {product['price']:,.0f}\n")
            
            if show_all_toko1:
                self.results_text.insert(tk.END, f"Lokasi: {product['location']}\n")
            else:
                self.results_text.insert(tk.END, f"Nama Toko: {product['source_store']}\n")
                
            self.results_text.insert(tk.END, "-" * 50 + "\n\n")
            
        status = "Display" if show_all_toko1 else "Comparison"
        self.results_text.insert(tk.END, f"\n{status} completed at {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")

    def save_results(self, toko1_name, toko2_list, unique_products, show_all_toko1=False):
        """Save comparison results to Excel file with user-selected location"""
        # Create default directory in Documents
        default_dir = os.path.join(str(Path.home()), "Documents", "CompareResult")
        os.makedirs(default_dir, exist_ok=True)

        # Default filename
        timestamp = datetime.now().strftime('%d%m%Y_%H%M%S')
        default_filename = f"Results - {timestamp}.xlsx"
        
        # Ask user for save location
        filename = filedialog.asksaveasfilename(
            initialdir=default_dir,
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save Comparison Results As"
        )
        
        if not filename:
            self.results_text.insert(tk.END, "\nFile save cancelled.")
            return
        
        # Create DataFrame with numbered index starting from 1
        df_data = []
        
        if show_all_toko1:
            # Show all products from Toko 1
            for idx, product in enumerate(unique_products, 1):
                nama_barang = product['name']
                kode_barang = self.extract_kode_barang(nama_barang)
                # Remove kode_barang from nama_barang if it exists
                if kode_barang != "0":
                    nama_barang = nama_barang[len(kode_barang):].strip(" -;")
                
                df_data.append({
                    'No': idx,
                    'Kode Barang': kode_barang,
                    'Nama Barang': nama_barang,
                    'Harga': f"Rp {product['price']:,.0f}",
                    'Lokasi': product['location']
                })
        else:
            # Show comparison results
            for idx, product in enumerate(unique_products, 1):
                nama_barang = product['name']
                kode_barang = self.extract_kode_barang(nama_barang)
                # Remove kode_barang from nama_barang if it exists
                if kode_barang != "0":
                    nama_barang = nama_barang[len(kode_barang):].strip(" -;")
                
                df_data.append({
                    'No': idx,
                    'Kode Barang': kode_barang,
                    'Nama Barang': nama_barang,
                    'Harga': f"Rp {product['price']:,.0f}",
                    'Nama Toko': product['source_store']
                })
                
        df = pd.DataFrame(df_data)
        
        # Add header information
        header_text = ['Comparison Results']
        header_text.append(f'Toko 1: {toko1_name}')
        if not show_all_toko1:
            comparison_stores = ", ".join(store[0] for store in toko2_list)
            header_text.append(f'Comparison Toko: {comparison_stores}')
        header_text.append(f'Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}')
        header_text.append('')  # Empty row before data
        
        header_df = pd.DataFrame(header_text)
        
        # Save to Excel
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Write header
            header_df.to_excel(writer, index=False, header=False, sheet_name='Sheet1')
            
            # Write main data starting from row 6
            df.to_excel(writer, index=False, header=True, sheet_name='Sheet1', startrow=5)
            
            # Get the workbook to apply styling
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            # Style the header
            for row in range(1, 5):
                cell = worksheet.cell(row=row, column=1)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left')
            
            # Style the table headers
            header_row = 6
            for col in range(1, 6):  # Adjusted for new column structure
                cell = worksheet.cell(row=header_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 5    # No
            worksheet.column_dimensions['B'].width = 15   # Kode Barang
            worksheet.column_dimensions['C'].width = 40   # Nama Barang
            worksheet.column_dimensions['D'].width = 15   # Harga
            worksheet.column_dimensions['E'].width = 30   # Nama Toko/Lokasi
        
        self.results_text.insert(tk.END, f"\nResults saved to {filename}")

def main():
    root = tk.Tk()
    app = CompareBarangApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
